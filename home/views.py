from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.http import JsonResponse, HttpResponse
from django.core import serializers
from django.db.models import Sum, F, FloatField, IntegerField, Case, When, Value, Q, Subquery, OuterRef
from django.utils import timezone
from django.db.models.functions import TruncDay, TruncMonth, Coalesce, Cast, TruncDate
import traceback
from django.contrib.auth import logout, login
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView
from .models import *
from django.contrib.auth.views import LoginView, PasswordChangeView, PasswordResetConfirmView, PasswordResetView
from admin_datta.forms import RegistrationForm, LoginForm, UserPasswordChangeForm, UserPasswordResetForm, UserSetPasswordForm
import json
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from import_export import resources
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from .forms import (
    CustomUserCreationForm, ProductForm, ExpensePurposeForm,
    ShopForm, ShopProductPriceForm, ExpenseForm, UserProfileForm
)
from django.urls import reverse

def is_admin_or_staff(user):
    return user.is_superuser or user.is_staff

def home(request):
    if request.user.is_superuser or request.user.groups.filter(name='manager').exists():
        return redirect('analytics')  # Redirect to analytics page
    return render(request, 'pages/home.html')

@login_required(login_url='/accounts/login/')
def submit_sales(request):
    if request.method == 'POST':
        try:
            # Parse JSON data from request body
            data = json.loads(request.body)
            
            # Get totals
            total_fedha = data.get('totalFedha')
            counted_money = data.get('countedMoney')
            difference = data.get('difference')
            products_data = data.get('products', [])

            # Create or update DailyCount
            today = timezone.now().date()
            daily_count, created = DailyCount.objects.get_or_create(
                date=today,
                shop=request.user.userprofile.shop,
                defaults={
                    'counted_amount': counted_money,
                    'user': request.user
                }
            )
            if not created:
                # Update the counted amount if already exists
                daily_count.counted_amount = counted_money
                daily_count.save()

            # Process each product
            for product_data in products_data:
                try:
                    product = Product.objects.get(id=product_data['productId'])
                    
                    # Create Sales record without counted_amount
                    Sales.objects.create(
                        product=product,
                        remainingJana=float(product_data['previousBalance']),
                        received=float(product_data['received']),
                        given=float(product_data['sold']),
                        total=float(product_data['total']),
                        sold=float(product_data['soldAmount']),
                        remaining=float(product_data['remainingBalance']),
                        spoiled=float(product_data['damaged']),
                        amount=float(product_data['cashAmount']),
                        difference=float(difference),
                        user=request.user,
                        shop=request.user.userprofile.shop
                    )
                    
                    # Update product's remaining value
                    product.remaining = float(product_data['remainingBalance'])
                    product.pokelewa_baridi = 0
                    product.pokelewa_moto = 0
                    product.pokelewa_mgando = 0
                    product.toka_baridi = 0
                    product.toka_moto = 0
                    product.toka_mgando = 0
                    product.save()
                    
                except Product.DoesNotExist:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Product with ID {product_data["productId"]} not found'
                    }, status=404)
                except Exception as e:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Error processing product: {str(e)}'
                    }, status=400)

            return JsonResponse({'status': 'success', 'message': 'Data submitted successfully!'})
            
        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    }, status=405)

@login_required(login_url='/accounts/login/')
@csrf_exempt
def expense_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            expenses = data.get('expenses', [])
            jumla_ya_matumizi = data.get('jumlaYaMatumizi', 0)

            for expense in expenses:
                purpose = expense.get('purpose')
                amount = expense.get('value', 0)

                Expenses.objects.create(
                    purpose=purpose,
                    cost=amount,
                    user_regst=request.user.username,
                    shop=request.user.userprofile.shop
                )

            return JsonResponse({'status': 'success', 'message': 'Expenses recorded successfully.'})
        except Exception as e:
            print(f"Error: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Method not allowed.'}, status=405)

@login_required(login_url='/accounts/login/')
def index(request):
    # Get today's date
    today = timezone.now().date()
    
    # Calculate daily sales
    daily_sales = Sales.objects.filter(
        created_at__date=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Calculate monthly sales
    first_day_of_month = today.replace(day=1)
    monthly_sales = Sales.objects.filter(
        created_at__date__gte=first_day_of_month,
        created_at__date__lte=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Calculate yearly sales
    first_day_of_year = today.replace(month=1, day=1)
    yearly_sales = Sales.objects.filter(
        created_at__date__gte=first_day_of_year,
        created_at__date__lte=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'segment': 'dashboard',
        'daily_sales': daily_sales,
        'monthly_sales': monthly_sales,
        'yearly_sales': yearly_sales,
    }
    return render(request, "pages/index.html", context)

@login_required(login_url='/accounts/login/')
def sales(request):
    # Get all products with their current remaining values
    products = Product.objects.all()  # Ensure this fetches the latest data
    expense_purpose = ExpensePurpose.objects.all()
    shops = Shop.objects.exclude(id=request.user.userprofile.shop.id)
    
    # Get suppliers from the supplier group associated with the user's shop
    suppliers = []
    try:
        supplier_group = Group.objects.get(name='Supplier')
        suppliers = User.objects.filter(
            groups=supplier_group,
            userprofile__shop=request.user.userprofile.shop
        )
    except Group.DoesNotExist:
        supplier_group = Group.objects.create(name='Supplier')
    
    context = {
        'segment': 'Sales',
        'products': products,  # Products now include the correct remaining value
        'expense_purpose': expense_purpose,
        'suppliers': suppliers,
        'shops': shops
    }
    return render(request, "pages/sales.html", context)

@login_required
def supplier(request):
  context = {
    'segment': 'supplier'
  }
  return render(request, "pages/supplier.html", context)

@login_required(login_url='/accounts/login/')
def table(request):
  context = {
    'segment': 'table'
  }
  return render(request, "pages/dynamic-tables.html", context)

@login_required(login_url='/accounts/login/')
def store(request):
    if request.method == 'POST':
        try:
            received_liters = request.POST.get('received_liters')
            given_liters = request.POST.get('given_liters')
            from_shop_id = request.POST.get('from_shop')
            to_shop_id = request.POST.get('to_shop')

            print(f"Received POST data: received_liters={received_liters}, given_liters={given_liters}, from_shop={from_shop_id}, to_shop={to_shop_id}")

            # Handle received liters
            if received_liters and from_shop_id and received_liters.strip():
                try:
                    received_liters = float(received_liters)
                    from_shop = Shop.objects.get(id=from_shop_id)
                    
                    print(f"Processing received liters: {received_liters} from shop {from_shop.name}")
                    
                    record = StoreTransfer.get_or_create_daily_record(
                        user=request.user,
                        from_shop=from_shop
                    )
                    
                    record.received_liters += received_liters
                    record.save()
                    print(f"Updated received record: {record}")
                except ValueError as e:
                    print(f"Error processing received liters: {e}")
                    return JsonResponse({'status': 'error', 'message': 'Invalid received liters value'})

            # Handle given liters
            if given_liters and to_shop_id and given_liters.strip():
                try:
                    given_liters = float(given_liters)
                    to_shop = Shop.objects.get(id=to_shop_id)
                    
                    print(f"Processing given liters: {given_liters} to shop {to_shop.name}")
                    
                    record = StoreTransfer.get_or_create_daily_record(
                        user=request.user,
                        to_shop=to_shop
                    )
                    
                    record.given_liters += given_liters
                    record.save()
                    print(f"Updated given record: {record}")
                except ValueError as e:
                    print(f"Error processing given liters: {e}")
                    return JsonResponse({'status': 'error', 'message': 'Invalid given liters value'})

            return redirect('store')
        except Exception as e:
            print(f"Error in store view: {str(e)}")
            print(traceback.format_exc())
            # toast.error(str(e))
            return HttpResponse({'status': 'error', 'message': str(e)})

    # Get filter parameters
    shop_id = request.GET.get('shop')
    user_id = request.GET.get('user')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Start with all transfers
    transfers = StoreTransfer.objects.all().order_by('-date_created')

    # Apply filters
    if shop_id:
        transfers = transfers.filter(models.Q(from_shop_id=shop_id) | models.Q(to_shop_id=shop_id))
    if user_id:
        transfers = transfers.filter(user_id=user_id)
    if date_from:
        transfers = transfers.filter(date_created__gte=date_from)
    if date_to:
        transfers = transfers.filter(date_created__lte=date_to)

    # Calculate totals
    total_received = transfers.aggregate(Sum('received_liters'))['received_liters__sum'] or 0
    total_given = transfers.aggregate(Sum('given_liters'))['given_liters__sum'] or 0

    # Get all shops and users for filters
    shops = Shop.objects.all()
    users = User.objects.filter(storetransfer__isnull=False).distinct()
    
    context = {
        'segment': 'store',
        'shops': shops,
        'transfers': transfers,
        'users': users,
        'total_received': total_received,
        'total_given': total_given,
        'selected_shop': shop_id,
        'selected_user': user_id,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
    }
    return render(request, "pages/store.html", context)

@login_required(login_url='/accounts/login/')
def analytics(request):
    # Get filter parameters
    shop_id = request.GET.get('shop')
    user_id = request.GET.get('user')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Start with all sales and expenses
    sales = Sales.objects.all()
    expenses = Expenses.objects.all()
    milk_stats = StoreTransfer.objects.all()

    # Apply filters to all querysets with proper timezone handling
    if shop_id:
        sales = sales.filter(shop_id=shop_id)
        expenses = expenses.filter(shop_id=shop_id)
        milk_stats = milk_stats.filter(models.Q(from_shop_id=shop_id) | models.Q(to_shop_id=shop_id))
    if user_id:
        sales = sales.filter(user_id=user_id)
        expenses = expenses.filter(user_regst=User.objects.get(id=user_id).username)
        milk_stats = milk_stats.filter(user_id=user_id)
    if date_from:
        try:
            date_from_aware = timezone.make_aware(
                datetime.strptime(date_from, '%Y-%m-%d')
            )
            sales = sales.filter(created_at__gte=date_from_aware)
            expenses = expenses.filter(date_created__gte=date_from_aware)
            milk_stats = milk_stats.filter(date_created__gte=date_from_aware)
        except (ValueError, TypeError):
            pass

    if date_to:
        try:
            # Add one day to include the entire end date
            date_to_aware = timezone.make_aware(
                datetime.strptime(date_to, '%Y-%m-%d')
            ) + timezone.timedelta(days=1)
            sales = sales.filter(created_at__lt=date_to_aware)
            expenses = expenses.filter(date_created__lt=date_to_aware)
            milk_stats = milk_stats.filter(date_created__lt=date_to_aware)
        except (ValueError, TypeError):
            pass

    # Calculate totals using the filtered querysets
    total_sales = sales.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expenses.aggregate(total=Sum('cost'))['total'] or 0
    
    # Calculate milk statistics from sales data
    total_received_hot = sales.filter(product__name='Maziwa Moto').aggregate(total=Sum('received'))['total'] or 0
    total_received_cold = sales.filter(product__name='Maziwa Baridi').aggregate(total=Sum('received'))['total'] or 0
    total_received_mgando = sales.filter(product__name='Maziwa Mgando').aggregate(total=Sum('received'))['total'] or 0
    total_received = total_received_hot + total_received_cold + total_received_mgando

    total_transferred_hot = sales.filter(product__name='Maziwa Moto').aggregate(total=Sum('given'))['total'] or 0
    total_transferred_cold = sales.filter(product__name='Maziwa Baridi').aggregate(total=Sum('given'))['total'] or 0
    total_transferred_mgando = sales.filter(product__name='Maziwa Mgando').aggregate(total=Sum('given'))['total'] or 0
    total_transferred = total_transferred_hot + total_transferred_cold + total_transferred_mgando

    # Calculate spoiled and difference
    total_spoiled = sales.aggregate(total=Sum('spoiled'))['total'] or 0
    total_difference = sales.aggregate(total=Sum('difference'))['total'] or 0

    # Calculate product-specific totals
    total_moto = sales.filter(product__name='Maziwa Moto').aggregate(total=Sum('amount'))['total'] or 0
    total_mgando = sales.filter(product__name='Maziwa Mgando').aggregate(total=Sum('amount'))['total'] or 0
    total_baridi = sales.filter(product__name='Maziwa Baridi').aggregate(total=Sum('amount'))['total'] or 0
    
    # Calculate vitafunwa total (excluding specific products)
    excluded_products = ['Maziwa Moto', 'Maziwa Mgando', 'Maziwa Baridi', 'Sukari', 'Yogati']
    total_vitafunwa = sales.exclude(product__name__in=excluded_products).aggregate(total=Sum('amount'))['total'] or 0

    # Get all shops and non-supplier users for the filter dropdowns
    shops = Shop.objects.all()
    supplier_group = Group.objects.get(name='Supplier')
    users = User.objects.exclude(groups=supplier_group).filter(sales__isnull=False).distinct()

    # Get filtered sales queryset
    sales = Sales.objects.all()
    if shop_id:
        sales = sales.filter(shop_id=shop_id)
    if user_id:
        sales = sales.filter(user_id=user_id)
    if date_from:
        try:
            date_from_aware = timezone.make_aware(datetime.strptime(date_from, '%Y-%m-%d'))
            sales = sales.filter(created_at__gte=date_from_aware)
        except (ValueError, TypeError):
            pass
    if date_to:
        try:
            date_to_aware = timezone.make_aware(datetime.strptime(date_to, '%Y-%m-%d')) + timezone.timedelta(days=1)
            sales = sales.filter(created_at__lt=date_to_aware)
        except (ValueError, TypeError):
            pass

    # Get daily records using filtered queryset
    daily_records = get_daily_records(sales)

    # Get supplier data aggregated by date
    supplier_data = SupplierData.objects.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        # POKELEWA aggregations
        wafugaji_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='supplier'), 
                 then='liter'),
            default=0,
            output_field=FloatField()
        )),
        jikoni_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='store'), 
                 then='liter'),
            default=0,
            output_field=FloatField()
        )),
        madukani_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='shop'), 
                 then='liter'),
            default=0,
            output_field=FloatField()
        )),
        
        # TOKA aggregations
        jikoni_transferred=Sum(Case(
            When(Q(action_type='toka') & Q(source_type='store'), 
                 then='liter'),
            default=0,
            output_field=FloatField()
        )),
        madukani_transferred=Sum(Case(
            When(Q(action_type='toka') & Q(source_type='shop'), 
                 then='liter'),
            default=0,
            output_field=FloatField()
        ))
    )

    # Convert to dictionary for easier lookup
    supplier_data_dict = {
        record['date']: record 
        for record in supplier_data
    }

    # Add supplier data to daily records
    for record in daily_records:
        date = record['created_at__date']
        record['supplier_data'] = supplier_data_dict.get(date, {
            'wafugaji_received': 0,
            'jikoni_received': 0,
            'madukani_received': 0,
            'jikoni_transferred': 0,
            'madukani_transferred': 0
        })

    # Calculate totals including supplier data
    totals = {
        'pokelewa_wafugaji': 0,
        'pokelewa_jikoni_mabichi': 0,
        'pokelewa_madukani_mabichi': 0,
        'pokelewa_jikoni_mgando': 0,
        'pokelewa_madukani_mgando': 0,
        'toka_jikoni_mabichi': 0,
        'toka_madukani_mabichi': 0,
        'toka_jikoni_mgando': 0,
        'toka_madukani_mgando': 0,
        'wafugaji': 0,
        'jikoni': 0,
        'madukani': 0,
        'uzwa_wafugaji': 0,
        'uzwa_moto': 0,
        'uzwa_mgando': 0,
        'haribika_mabichi': 0,
        'haribika_moto': 0,
        'haribika_mgando': 0,
        'baki_mabichi': 0,
        'baki_moto': 0,
        'baki_mgando': 0,
        'chapati': 0,
        'hcafe': 0,
        'maandazi': 0,
        'sconzi': 0,
        'vitumbua': 0,
        'mikate': 0,
        'mayai': 0,
        'bagia': 0,
        'mauzo': 0,
        'vitafunwa_total': 0,
        'halisi': 0,
        'expenses': 0,
        'cash': 0,
        'baridi': 0,
        'moto': 0,
        'mgando': 0,
        'difference': 0,
        'supplier_data': {
            'wafugaji_received': 0,
            'jikoni_received': 0,
            'madukani_received': 0,
            'jikoni_transferred': 0,
            'madukani_transferred': 0
        },
        'total_vitafunwa': total_vitafunwa,
    }
    
    for record in daily_records:
        for key in totals.keys():
            if key == 'supplier_data':
                # Handle supplier_data separately
                for supplier_key in totals['supplier_data'].keys():
                    try:
                        value = float(record.get('supplier_data', {}).get(supplier_key, 0))
                    except (TypeError, ValueError):
                        value = 0.0
                    totals['supplier_data'][supplier_key] += value
            else:
                # Handle other totals as before
                try:
                    value = float(record.get(key, 0))
                except (TypeError, ValueError):
                    value = 0.0
                totals[key] += value

    # Calculate total cash (total sales - expenses)
    total_cash = total_sales - total_expenses if total_sales and total_expenses else 0

    # Calculate total sold milk
    total_sold_hot = Sales.objects.filter(
        product__name='Maziwa Moto'
    ).aggregate(
        total=Coalesce(Sum('sold'), Value(0), output_field=FloatField())
    )['total']

    total_sold_cold = Sales.objects.filter(
        product__name='Maziwa Baridi'
    ).aggregate(
        total=Coalesce(Sum('sold'), Value(0), output_field=FloatField())
    )['total']

    total_sold_mgando = Sales.objects.filter(
        product__name='Maziwa Mgando'
    ).aggregate(
        total=Coalesce(Sum('sold'), Value(0), output_field=FloatField())
    )['total']

    total_sold_milk = total_sold_hot + total_sold_cold + total_sold_mgando

    # Calculate spoiled milk by type
    total_spoiled_hot = Sales.objects.filter(
        product__name='Maziwa Moto'
    ).aggregate(
        total=Coalesce(Sum('spoiled'), Value(0), output_field=FloatField())
    )['total']

    total_spoiled_cold = Sales.objects.filter(
        product__name='Maziwa Baridi'
    ).aggregate(
        total=Coalesce(Sum('spoiled'), Value(0), output_field=FloatField())
    )['total']

    total_spoiled_mgando = Sales.objects.filter(
        product__name='Maziwa Mgando'
    ).aggregate(
        total=Coalesce(Sum('spoiled'), Value(0), output_field=FloatField())
    )['total']

    context = {
        'segment': 'analytics',
        'sales': sales,
        'total_sales': total_sales,
        'total_expenses': total_expenses,
        'total_cash': total_cash,
        'total_received': total_received,
        'total_received_hot': total_received_hot,
        'total_received_cold': total_received_cold,
        'total_received_mgando': total_received_mgando,
        'total_transferred': total_transferred,
        'total_transferred_hot': total_transferred_hot,
        'total_transferred_cold': total_transferred_cold,
        'total_transferred_mgando': total_transferred_mgando,
        'total_spoiled': total_spoiled,
        'total_difference': total_difference,
        'total_moto': total_moto,
        'total_mgando': total_mgando,
        'total_baridi': total_baridi,
        'total_vitafunwa': total_vitafunwa,
        'shops': shops,
        'users': users,
        'daily_records': daily_records,
        'totals': totals,
        'total_sold_milk': total_sold_milk,
        'total_sold_hot': total_sold_hot,
        'total_sold_cold': total_sold_cold,
        'total_sold_mgando': total_sold_mgando,
        'total_spoiled_hot': total_spoiled_hot,
        'total_spoiled_cold': total_spoiled_cold,
        'total_spoiled_mgando': total_spoiled_mgando,
    }
    
    return render(request, 'pages/analytics.html', context)

# Authentication
class UserLoginView(LoginView):
  template_name = 'accounts/login.html'
  form_class = LoginForm

@login_required
@user_passes_test(is_admin_or_staff)
def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        profile_form = UserProfileForm(request.POST)
        if form.is_valid() and profile_form.is_valid():
            user = form.save()
            group = form.cleaned_data['group']
            user.groups.add(group)
            
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()
            
            messages.success(request, 'User created successfully!')
            return redirect('admin:auth_user_changelist')
    else:
        form = CustomUserCreationForm()
        profile_form = UserProfileForm()

    context = {
        'form': form,
        'profile_form': profile_form,
        'product_form': ProductForm(),
        'shop_form': ShopForm(),
        'expense_form': ExpensePurposeForm(),
        'price_form': ShopProductPriceForm(),
    }
    return render(request, 'accounts/register.html', context)

def logout_view(request):
  logout(request)
  return redirect('/accounts/login/')

class UserPasswordResetView(PasswordResetView):
  template_name = 'accounts/password_reset.html'
  form_class = UserPasswordResetForm

class UserPasswordResetConfirmView(PasswordResetConfirmView):
  template_name = 'accounts/password_reset_confirm.html'
  form_class = UserSetPasswordForm

class UserPasswordChangeView(PasswordChangeView):
  template_name = 'accounts/password_change.html'
  form_class = UserPasswordChangeForm

def register_view(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            shop_id = request.POST.get('shop')
            shop = Shop.objects.get(id=shop_id)
            UserProfile.objects.create(user=user, shop=shop)
            return redirect('login')
    else:
        form = RegistrationForm()
    
    shops = Shop.objects.all()
    return render(request, 'accounts/register.html', {'form': form, 'shops': shops})

@login_required
def shop_report(request, shop_id):
    shop = Shop.objects.get(id=shop_id)
    sales = Sales.objects.filter(shop=shop)
    # Add your report logic here
    return render(request, 'reports/shop_report.html', {'sales': sales, 'shop': shop})

@login_required
def user_report(request, user_id):
    user_sales = Sales.objects.filter(user_id=user_id)
    # Add your report logic here
    return render(request, 'reports/user_report.html', {'sales': user_sales})

@login_required(login_url='/accounts/login/')
def submit_supplier_data(request):
    if request.method == 'POST':
        try:
            source_type = request.POST.get('source_type')
            source_id = request.POST.get('source_id')
            source_name = request.POST.get('source_name')
            action_type = request.POST.get('action_type')  
            liter = float(request.POST.get('liter', 0))
            milk_type = request.POST.get('milk_type')
            action_type = request.POST.get('action_type')

            # Save the data to the database for analytics
            SupplierData.objects.create(
                source_type=source_type,
                source_id=source_id,
                source_name=source_name, 
                action_type=action_type,
                liter=liter,
                milk_type=milk_type,
                staff=request.user
            )

            # Update the corresponding product's pokelewa and toka fields
            try:
                product = Product.objects.get(name=milk_type)
                
                if action_type == 'pokea':
                    if milk_type == 'Maziwa Moto':
                        product.pokelewa_moto += liter
                    elif milk_type == 'Maziwa Baridi':
                        product.pokelewa_baridi += liter
                    elif milk_type == 'Maziwa Mgando':
                        product.pokelewa_mgando += liter
                
                elif action_type == 'toka':
                    if milk_type == 'Maziwa Moto':
                        product.pokelewa_moto -= liter
                        product.toka_moto += liter
                    elif milk_type == 'Maziwa Baridi':
                        product.pokelewa_baridi -= liter
                        product.toka_baridi += liter
                    elif milk_type == 'Maziwa Mgando':
                        product.pokelewa_mgando -= liter
                        product.toka_mgando += liter
                
                product.save()
            except Product.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)

            return JsonResponse({
                'status': 'success', 
                'message': 'Data saved successfully!'
            })
        except Exception as e:
            print(f"Error: {str(e)}")  # Add this for debugging
            return JsonResponse({
                'status': 'error', 
                'message': str(e)
            }, status=500)
    else:
        return JsonResponse({
            'status': 'error', 
            'message': 'Invalid request method'
        }, status=405)

@login_required(login_url='/accounts/login/')
def supplier_view(request):
    # Get user's group
    is_admin = request.user.groups.filter(name='Admin').exists()
    
    # Base queryset
    queryset = SupplierData.objects.all()
    
    # Filter by user if not admin
    if not is_admin:
        queryset = queryset.filter(staff=request.user)
    
    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    period = request.GET.get('period', 'month')
    
    today = timezone.now()
    
    # Apply date filters if provided
    if start_date:
        start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
        start_date = timezone.make_aware(start_date)
        queryset = queryset.filter(created_at__gte=start_date)
    else:
        if period == 'week':
            start_date = today - timezone.timedelta(days=7)
        elif period == 'year':
            start_date = today - timezone.timedelta(days=365)
        else:  # month
            start_date = today - timezone.timedelta(days=30)
        queryset = queryset.filter(created_at__gte=start_date)
    
    if end_date:
        end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')
        end_date = timezone.make_aware(end_date)
        end_date = end_date + timezone.timedelta(days=1)  # Include the end date
        queryset = queryset.filter(created_at__lt=end_date)
    else:
        end_date = today
    
    # Set current period text
    if start_date and end_date:
        current_period = f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
    elif period == 'week':
        current_period = 'This Week'
    elif period == 'year':
        current_period = 'This Year'
    else:
        current_period = 'This Month'
    
    # Calculate statistics
    period_data = queryset
    total_liters = period_data.aggregate(total=models.Sum('liter'))['total'] or 0
    
    # Calculate growth (comparing with previous period)
    time_diff = end_date - start_date
    previous_start = start_date - time_diff
    previous_data = SupplierData.objects.filter(
        created_at__gte=previous_start,
        created_at__lt=start_date
    )
    if not is_admin:
        previous_data = previous_data.filter(staff=request.user)
    
    previous_liters = previous_data.aggregate(total=models.Sum('liter'))['total'] or 0
    liter_growth = ((total_liters - previous_liters) / previous_liters * 100) if previous_liters else 0
    
    # Calculate revenue (assuming a fixed price per liter)
    price_per_liter = 1000  # Set your price here
    total_revenue = total_liters * price_per_liter
    previous_revenue = previous_liters * price_per_liter
    revenue_growth = ((total_revenue - previous_revenue) / previous_revenue * 100) if previous_revenue else 0
    
    # Get supply trend data
    if (end_date - start_date).days <= 31:
        trend_data = queryset.values('source_name').annotate(
            date=TruncDay('created_at')
        ).values('date', 'source_name').annotate(
            total=models.Sum('liter')
        ).order_by('date', 'source_name')
    else:
        trend_data = queryset.values('source_name').annotate(
            date=TruncMonth('created_at')
        ).values('date', 'source_name').annotate(
            total=models.Sum('liter')
        ).order_by('date', 'source_name')

    # Process trend data to create series for each supplier
    trend_by_supplier = {}
    trend_labels = []
    seen_dates = set()

    for entry in trend_data:
        date_str = entry['date'].strftime('%d %b' if (end_date - start_date).days <= 31 else '%b %Y')
        if date_str not in seen_dates:
            trend_labels.append(date_str)
            seen_dates.add(date_str)
        
        supplier = entry['source_name']
        if supplier not in trend_by_supplier:
            trend_by_supplier[supplier] = [0] * len(trend_labels)
        
        trend_by_supplier[supplier][len(trend_labels) - 1] = float(entry['total'])

    # Convert to series format for ApexCharts
    supply_trend_series = [
        {
            'name': supplier,
            'data': values
        }
        for supplier, values in trend_by_supplier.items()
    ]

    # Get top suppliers data
    top_suppliers = queryset.values('source_name').annotate(
        total=models.Sum('liter')
    ).order_by('-total')[:5]
    
    top_suppliers_labels = [s['source_name'] for s in top_suppliers]
    top_suppliers_data = [float(s['total']) for s in top_suppliers]

    # Get months for filter
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'),
        (4, 'April'), (5, 'May'), (6, 'June'),
        (7, 'July'), (8, 'August'), (9, 'September'),
        (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Add revenue to queryset
    supplier_data = period_data.order_by('-created_at')
    for record in supplier_data:
        record.revenue = record.liter * price_per_liter
    
    context = {
        'total_liters': total_liters,
        'liter_growth': round(liter_growth, 1),
        'total_revenue': total_revenue,
        'revenue_growth': round(revenue_growth, 1),
        'total_suppliers': period_data.values('source_name').distinct().count(),
        'current_period': current_period,
        'supply_trend_series': supply_trend_series,
        'supply_trend_labels': trend_labels,
        'top_suppliers_data': top_suppliers_data,
        'top_suppliers_labels': top_suppliers_labels,
        'months': months,
        'supplier_data': supplier_data,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'pages/supplier.html', context)

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            selected_shop_id = request.POST.get('selected_shop')
            remember = form.cleaned_data.get('remember', False)

            # Login the user first
            login(request, user)

            # First check if user is superuser or staff
            if user.is_superuser or user.is_staff:
                return redirect('analytics')  # Changed from /admin/ to analytics

            # Then handle regular user groups
            if user.groups.filter(name='Manager').exists():
                if not selected_shop_id:
                    messages.error(request, 'Please select a shop to login.')
                    logout(request)  # Logout the user if no shop selected
                    return render(request, 'accounts/login.html', {'form': form})
                
                try:
                    # Get the selected shop
                    selected_shop = Shop.objects.get(id=selected_shop_id)
                    
                    # Update the user's profile with the selected shop
                    user_profile = UserProfile.objects.get(user=user)
                    user_profile.shop = selected_shop
                    user_profile.save()
                    
                    # Store selected shop in session
                    request.session['selected_shop'] = selected_shop_id
                    
                    return redirect('sales')  # Redirect managers to sales page
                except Shop.DoesNotExist:
                    messages.error(request, 'Invalid shop selection.')
                    logout(request)
                    return render(request, 'accounts/login.html', {'form': form})
                
            elif user.groups.filter(name='Storekeeper').exists():
                return redirect('store')  # Redirect storekeepers to store page

            elif user.groups.filter(name='Supplier').exists():
                return redirect('supplier')  # Redirect suppliers to supplier page

            # Handle session expiry for remember me
            if not remember:
                request.session.set_expiry(0)
                
            # Default redirect
            return redirect('index')

    else:
        form = LoginForm()
    
    return render(request, 'accounts/login.html', {'form': form})

def check_user_group(request, username):
    try:
        user = User.objects.get(username=username)
        # Check if user is superuser or staff first
        if user.is_superuser or user.is_staff:
            return JsonResponse({
                'status': 'success',
                'is_manager': False  # Skip shop selection for admin users
            }, safe=False)
            
        is_manager = user.groups.filter(name='Manager').exists()
        return JsonResponse({
            'status': 'success',
            'is_manager': is_manager
        }, safe=False)
    except ObjectDoesNotExist:
        return JsonResponse({
            'status': 'error',
            'is_manager': False,
            'message': 'User not found'
        }, safe=False)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'is_manager': False,
            'message': str(e)
        }, safe=False)

def get_shops(request):
    """Return all shops as JSON for the shop selection dropdown"""
    shops = Shop.objects.all()
    shops_data = [{'id': shop.id, 'name': shop.name} for shop in shops]
    return JsonResponse({'shops': shops_data})

# Create a resource for Sales model
class SalesResource(resources.ModelResource):
    class Meta:
        model = Sales

@login_required(login_url='/accounts/login/')
def export_sales_excel(request):
    sales_resource = SalesResource()
    dataset = sales_resource.export()
    response = HttpResponse(dataset.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales.xlsx"'
    return response

@login_required(login_url='/accounts/login/')
def export_sales_pdf(request):
    sales = Sales.objects.all()
    html = render_to_string('pages/sales_pdf.html', {'sales': sales})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@login_required(login_url='/accounts/login/')
def export_daily_analysis_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="daily_analysis.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Daily Analysis"

    # Get daily records first
    daily_records = get_daily_records()
    
    # Get supplier data
    supplier_data = SupplierData.objects.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        wafugaji_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='supplier'), then='liter'),
            default=0,
            output_field=FloatField()
        )),
        jikoni_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='store'), then='liter'),
            default=0,
            output_field=FloatField()
        )),
        madukani_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='shop'), then='liter'),
            default=0,
            output_field=FloatField()
        )),
        jikoni_transferred=Sum(Case(
            When(Q(action_type='toka') & Q(source_type='store'), then='liter'),
            default=0,
            output_field=FloatField()
        )),
        madukani_transferred=Sum(Case(
            When(Q(action_type='toka') & Q(source_type='shop'), then='liter'),
            default=0,
            output_field=FloatField()
        ))
    )

    # Convert to dictionary for easier lookup
    supplier_data_dict = {record['date']: record for record in supplier_data}

    # Add supplier data to daily records
    for record in daily_records:
        date = record['created_at__date']
        record['supplier_data'] = supplier_data_dict.get(date, {
            'wafugaji_received': 0,
            'jikoni_received': 0,
            'madukani_received': 0,
            'jikoni_transferred': 0,
            'madukani_transferred': 0
        })

    # Define colors matching web view
    colors = {
        'light': 'F8F9FA',
        'info': 'CFE2FF',
        'success': 'D1E7DD',
        'warning': 'FFF3CD',
        'danger': 'F8D7DA',
        'primary': 'CCE5FF',
        'secondary': 'E2E3E5'
    }

    # Add title and date (rows 1-2)
    title = "ISANGA DAIRY - DETAILED DAILY ANALYSIS"
    worksheet.merge_cells('A1:AJ1')
    title_cell = worksheet['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    date_str = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    worksheet.merge_cells('A2:AJ2')
    date_cell = worksheet['A2']
    date_cell.value = date_str
    date_cell.font = Font(italic=True)
    date_cell.alignment = Alignment(horizontal='center', vertical='center')

    # First header row (row 4) - Main Categories
    worksheet.append([''] * 36)  # Empty row 3
    first_row = ['Tarehe']
    first_row_merges = [
        ('B4:F4', 'POKELEWA', colors['info']),
        ('G4:K4', 'TOKA', colors['success']),
        ('L4:M4', 'UZWA', colors['warning']),
        ('N4:P4', 'HARIBIKA', colors['danger']),
        ('Q4:S4', 'BAKI', colors['primary']),
        ('T4:AA4', 'VITAFUNWA', colors['secondary']),
        ('AB4:AJ4', 'FEDHA', colors['light'])
    ]
    worksheet.append(first_row + [''] * 35)  # Add empty cells for merging

    # Apply merges and styles for first row
    for merge_range, text, color in first_row_merges:
        worksheet.merge_cells(merge_range)
        merged_cell = worksheet[merge_range.split(':')[0]]
        merged_cell.value = text
        merged_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = Font(bold=True)

    # Second header row (row 5) - Subcategories
    second_row_merges = [
        ('B5:D5', 'MABICHI', colors['info']),
        ('E5:F5', 'MGANDO', colors['info']),
        ('G5:H5', 'MABICHI', colors['success']),
        ('I5:K5', 'MGANDO', colors['success']),
        ('L5:M5', 'MAZIWA', colors['warning']),
        ('N5:P5', 'MAZIWA', colors['danger']),
        ('Q5:S5', 'MAZIWA', colors['primary']),
        ('T5:AA5', 'BIDHAA', colors['secondary']),
        ('AB5:AJ5', 'MAPATO/MATUMIZI', colors['light'])
    ]
    worksheet.append(['Tarehe'] + [''] * 35)

    # Apply merges and styles for second row
    for merge_range, text, color in second_row_merges:
        worksheet.merge_cells(merge_range)
        merged_cell = worksheet[merge_range.split(':')[0]]
        merged_cell.value = text
        merged_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = Font(bold=True)

    # Third header row (row 6) - Detailed headers
    third_row_headers = [
        'Tarehe',
        # POKELEWA
        'WAFUGAJI', 'JIKONI', 'MADUKANI', 'JIKONI', 'MADUKANI',
        # TOKA
        'JIKONI', 'MADUKANI', 'JIKONI', 'MADUKANI', 'WAFUGAJI',
        # UZWA
        'MOTO', 'MGANDO',
        # HARIBIKA
        'MABICHI', 'MOTO', 'MGANDO',
        # BAKI
        'MABICHI', 'MOTO', 'MGANDO',
        # VITAFUNWA
        'CHAPATI', 'H.CAFE', 'MAANDAZI', 'SCONZI', 'VITUMBUA', 'MIKATE', 'MAYAI', 'BAGIA',
        # FEDHA
        'MAUZO', 'HALISI', 'MATUMIZI', 'CASH', 'BARIDI', 'MOTO', 'MGANDO', 'VITAFUNWA', 'TOFAUTI'
    ]
    worksheet.append(third_row_headers)

    # Style third row
    for cell in worksheet[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Set background color based on column group
        col_idx = cell.column - 1
        if col_idx < 1:  # Tarehe
            color = colors['light']
        elif col_idx < 6:  # POKELEWA
            color = colors['info']
        elif col_idx < 11:  # TOKA
            color = colors['success']
        elif col_idx < 13:  # UZWA
            color = colors['warning']
        elif col_idx < 16:  # HARIBIKA
            color = colors['danger']
        elif col_idx < 19:  # BAKI
            color = colors['primary']
        elif col_idx < 27:  # VITAFUNWA
            color = colors['secondary']
        else:  # FEDHA
            color = colors['light']
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    # Add data rows with styling
    row_num = 7
    for record in daily_records:
        row = [
            record['created_at__date'].strftime('%d/%m/%Y'),
            # POKELEWA
            record['supplier_data']['wafugaji_received'],
            record['supplier_data']['jikoni_received'],
            record['supplier_data']['madukani_received'],
            record['pokelewa_jikoni_mgando'],
            record['pokelewa_madukani_mgando'],
            # TOKA
            record['supplier_data']['jikoni_transferred'],
            record['supplier_data']['madukani_transferred'],
            record['toka_jikoni_mgando'],
            record['toka_madukani_mgando'],
            record['uzwa_wafugaji'],
            # UZWA
            record['uzwa_moto'],
            record['uzwa_mgando'],
            # HARIBIKA
            record['haribika_mabichi'],
            record['haribika_moto'],
            record['haribika_mgando'],
            # BAKI
            record['baki_mabichi'],
            record['baki_moto'],
            record['baki_mgando'],
            # VITAFUNWA
            record['chapati'],
            record['hcafe'],
            record['maandazi'],
            record['sconzi'],
            record['vitumbua'],
            record['mikate'],
            record['mayai'],
            record['bagia'],
            # FEDHA
            record['mauzo'],
            record['halisi'],
            record['expenses'],
            record['cash'],
            record['baridi'],
            record['moto'],
            record['mgando'],
            record['vitafunwa_total'],
            record['difference']
        ]
        worksheet.append(row)
        
        # Style data cells
        for cell in worksheet[row_num]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Format numbers
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        row_num += 1

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            if isinstance(cell, MergedCell):
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Freeze panes to keep headers visible
    worksheet.freeze_panes = 'A7'

    workbook.save(response)
    return response

@login_required(login_url='/accounts/login/')
def export_daily_analysis_pdf(request):
    daily_records = get_daily_records()
    context = {
        'daily_records': daily_records,
        'title': 'Daily Analysis Report'
    }
    html = render_to_string('pages/daily_analysis_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="daily_analysis.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

def get_daily_records(sales_queryset=None):
    if sales_queryset is None:
        sales_queryset = Sales.objects.all()

    # First, let's print out some debug information
    print("Debug: Checking sales records before aggregation")
    debug_data = sales_queryset.values(
        'created_at__date', 'shop__name', 'user__username'
    ).order_by('created_at__date')
    for record in debug_data:
        print(f"Date: {record['created_at__date']}, Shop: {record['shop__name']}, User: {record['user__username']}")

    # Get daily records with proper date grouping
    daily_data = sales_queryset.annotate(
        date=TruncDate('created_at')
    ).values(
        'date'  # Only group by date, ignore shop and user
    ).annotate(
        # POKELEWA
        pokelewa_jikoni_mgando=Coalesce(
            Sum('received', filter=Q(product__name='Maziwa Mgando')), 
            Value(0), 
            output_field=FloatField()
        ),
        pokelewa_madukani_mgando=Coalesce(
            Sum('received', filter=Q(product__name='Maziwa Mgando')), 
            Value(0), 
            output_field=FloatField()
        ),

        # TOKA
        toka_jikoni_mgando=Coalesce(
            Sum('given', filter=Q(product__name='Maziwa Mgando')), 
            Value(0), 
            output_field=FloatField()
        ),
        toka_madukani_mgando=Coalesce(
            Sum('given', filter=Q(product__name='Maziwa Mgando')), 
            Value(0), 
            output_field=FloatField()
        ),

        # UZWA (using sold values)
        uzwa_wafugaji=Coalesce(
            Sum('sold', filter=Q(product__name='Wafugaji')), 
            Value(0), 
            output_field=FloatField()
        ),
        uzwa_moto=Coalesce(
            Sum('sold', filter=Q(product__name='Maziwa Moto')), 
            Value(0), 
            output_field=FloatField()
        ),
        uzwa_mgando=Coalesce(
            Sum('sold', filter=Q(product__name='Maziwa Mgando')), 
            Value(0), 
            output_field=FloatField()
        ),

        # HARIBIKA (Spoiled)
        haribika_mabichi=Coalesce(
            Sum('spoiled', filter=Q(product__name='Maziwa Baridi')), 
            Value(0), 
            output_field=FloatField()
        ),
        haribika_moto=Coalesce(
            Sum('spoiled', filter=Q(product__name='Maziwa Moto')), 
            Value(0), 
            output_field=FloatField()
        ),
        haribika_mgando=Coalesce(
            Sum('spoiled', filter=Q(product__name='Maziwa Mgando')), 
            Value(0), 
            output_field=FloatField()
        ),

        # BAKI (Remaining)
        baki_mabichi=Coalesce(
            Sum('remaining', filter=Q(product__name='Maziwa Baridi')), 
            Value(0), 
            output_field=FloatField()
        ),
        baki_moto=Coalesce(
            Sum('remaining', filter=Q(product__name='Maziwa Moto')), 
            Value(0), 
            output_field=FloatField()
        ),
        baki_mgando=Coalesce(
            Sum('remaining', filter=Q(product__name='Maziwa Mgando')), 
            Value(0), 
            output_field=FloatField()
        ),

        # VITAFUNWA (Products) - using sold values
        chapati=Coalesce(
            Sum('sold', filter=Q(product__name='Chapati')), 
            Value(0), 
            output_field=FloatField()
        ),
        hcafe=Coalesce(
            Sum('sold', filter=Q(product__name='H.Cafe')), 
            Value(0), 
            output_field=FloatField()
        ),
        maandazi=Coalesce(
            Sum('sold', filter=Q(product__name='Maandazi')), 
            Value(0), 
            output_field=FloatField()
        ),
        sconzi=Coalesce(
            Sum('sold', filter=Q(product__name='Sconzi')), 
            Value(0), 
            output_field=FloatField()
        ),
        vitumbua=Coalesce(
            Sum('sold', filter=Q(product__name='Vitumbua')), 
            Value(0), 
            output_field=FloatField()
        ),
        mikate=Coalesce(
            Sum('sold', filter=Q(product__name='Mikate')), 
            Value(0), 
            output_field=FloatField()
        ),
        mayai=Coalesce(
            Sum('sold', filter=Q(product__name='Mayai')), 
            Value(0), 
            output_field=FloatField()
        ),
        bagia=Coalesce(
            Sum('sold', filter=Q(product__name='Bagia')), 
            Value(0), 
            output_field=FloatField()
        ),

        # FEDHA
        mauzo=Coalesce(Sum('amount'), Value(0), output_field=FloatField()),
        vitafunwa_total=Coalesce(
            Sum('amount', filter=Q(
                product__name__in=[
                    'Chapati', 'H.Cafe', 'Maandazi', 'Sconzi',
                    'Vitumbua', 'Mikate', 'Mayai', 'Bagia'
                ]
            )),
            Value(0),
            output_field=FloatField()
        ),
        halisi=Coalesce(
            Subquery(
                DailyCount.objects.filter(
                    date=OuterRef('date')
                ).values('date').annotate(
                    total=Sum('counted_amount')
                ).values('total')
            ),
            Value(0),
            output_field=FloatField()
        ),
        expenses=Coalesce(
            Subquery(
                Expenses.objects.filter(
                    date_created=OuterRef('date')
                ).values('date_created').annotate(
                    total=Sum('cost')
                ).values('total')
            ),
            Value(0),
            output_field=FloatField()
        ),
        cash=F('halisi') - F('expenses'),
        baridi=Coalesce(Sum('amount', filter=Q(product__name='Maziwa Baridi')), Value(0), output_field=FloatField()),
        moto=Coalesce(Sum('amount', filter=Q(product__name='Maziwa Moto')), Value(0), output_field=FloatField()),
        mgando=Coalesce(Sum('amount', filter=Q(product__name='Maziwa Mgando')), Value(0), output_field=FloatField()),
        difference=Coalesce(Sum('difference'), Value(0), output_field=FloatField())
    ).order_by('date')

    # Get supplier data for these dates
    supplier_data = SupplierData.objects.filter(
        created_at__date__in=daily_data.values('date')
    ).annotate(
        date=TruncDate('created_at')
    ).values(
        'date'
    ).annotate(
        wafugaji_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='supplier'), then='liter'),
            default=Value(0),
            output_field=FloatField()
        )),
        jikoni_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='store'), then='liter'),
            default=Value(0),
            output_field=FloatField()
        )),
        madukani_received=Sum(Case(
            When(Q(action_type='pokea') & Q(source_type='shop'), then='liter'),
            default=Value(0),
            output_field=FloatField()
        )),
        jikoni_transferred=Sum(Case(
            When(Q(action_type='toka') & Q(source_type='store'), then='liter'),
            default=Value(0),
            output_field=FloatField()
        )),
        madukani_transferred=Sum(Case(
            When(Q(action_type='toka') & Q(source_type='shop'), then='liter'),
            default=Value(0),
            output_field=FloatField()
        ))
    )

    # Convert supplier data to dictionary for easy lookup
    supplier_data_dict = {
        record['date']: {
            'wafugaji_received': record['wafugaji_received'],
            'jikoni_received': record['jikoni_received'],
            'madukani_received': record['madukani_received'],
            'jikoni_transferred': record['jikoni_transferred'],
            'madukani_transferred': record['madukani_transferred']
        }
        for record in supplier_data
    }

    # Convert daily_data to list and add supplier data
    daily_data_list = list(daily_data)
    for record in daily_data_list:
        date = record['date']
        supplier_info = supplier_data_dict.get(date, {
            'wafugaji_received': 0,
            'jikoni_received': 0,
            'madukani_received': 0,
            'jikoni_transferred': 0,
            'madukani_transferred': 0
        })
        record['supplier_data'] = supplier_info
        record['created_at__date'] = date

    # Print debug information after aggregation
    print("\nDebug: Checking aggregated records")
    for record in daily_data_list:
        print(f"Date: {record['date']}, Total Sales: {record['mauzo']}")

    return daily_data_list

@login_required
@user_passes_test(is_admin_or_staff)
def register_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            group = form.cleaned_data['group']
            shop = form.cleaned_data['shop']
            user.groups.add(group)
            UserProfile.objects.create(user=user, shop=shop)
            messages.success(request, 'User created successfully!')
            return redirect('user_list')
    else:
        form = CustomUserCreationForm()
    return render(request, 'accounts/register.html', {'form': form})

@login_required
@user_passes_test(is_admin_or_staff)
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product created successfully!')
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'accounts/model_form.html', {
        'form': form,
        'form_title': 'Create Product'
    })

@login_required
@user_passes_test(is_admin_or_staff)
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated successfully!')
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'accounts/model_form.html', {
        'form': form,
        'form_title': 'Edit Product'
    })

@login_required
@user_passes_test(is_admin_or_staff)
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully!')
        return redirect('product_list')
    return render(request, 'accounts/confirm_delete.html', {
        'object': product,
        'title': 'Delete Product'
    })

@login_required
@user_passes_test(is_admin_or_staff)
def product_list(request):
    products = Product.objects.all()
    items = [{
        'fields': [p.name, p.default_price],
        'edit_url': reverse('product_edit', args=[p.id]),
        'delete_url': reverse('product_delete', args=[p.id])
    } for p in products]
    
    return render(request, 'accounts/model_list.html', {
        'items': items,
        'headers': ['Name', 'Default Price'],
        'list_title': 'Products',
        'create_url': reverse('product_create')
    })

@login_required
@user_passes_test(is_admin_or_staff)
def expense_purpose_list(request):
    purposes = ExpensePurpose.objects.all()
    items = [{
        'fields': [p.name],
        'edit_url': reverse('expense_purpose_edit', args=[p.id]),
        'delete_url': reverse('expense_purpose_delete', args=[p.id])
    } for p in purposes]
    
    return render(request, 'accounts/model_list.html', {
        'items': items,
        'headers': ['Name'],
        'list_title': 'Expense Purposes',
        'create_url': reverse('expense_purpose_create')
    })

@login_required
@user_passes_test(is_admin_or_staff)
def shop_list(request):
    shops = Shop.objects.all()
    items = [{
        'fields': [s.name, s.location],
        'edit_url': reverse('shop_edit', args=[s.id]),
        'delete_url': reverse('shop_delete', args=[s.id])
    } for s in shops]
    
    return render(request, 'accounts/model_list.html', {
        'items': items,
        'headers': ['Name', 'Location'],
        'list_title': 'Shops',
        'create_url': reverse('shop_create')
    })

@login_required
@user_passes_test(is_admin_or_staff)
def shop_price_list(request):
    prices = ShopProductPrice.objects.all()
    items = [{
        'fields': [p.shop.name, p.product.name, p.price],
        'edit_url': reverse('shop_price_edit', args=[p.id]),
        'delete_url': reverse('shop_price_delete', args=[p.id])
    } for p in prices]
    
    return render(request, 'accounts/model_list.html', {
        'items': items,
        'headers': ['Shop', 'Product', 'Price'],
        'list_title': 'Shop Prices',
        'create_url': reverse('shop_price_create')
    })

@login_required
@user_passes_test(is_admin_or_staff)
def expense_purpose_create(request):
    if request.method == 'POST':
        form = ExpensePurposeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense Purpose created successfully!')
            return redirect('expense_purpose_list')
    else:
        form = ExpensePurposeForm()
    return render(request, 'accounts/model_form.html', {
        'form': form,
        'form_title': 'Create Expense Purpose'
    })

@login_required
@user_passes_test(is_admin_or_staff)
def shop_create(request):
    if request.method == 'POST':
        form = ShopForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Shop created successfully!')
            return redirect('shop_list')
    else:
        form = ShopForm()
    return render(request, 'accounts/model_form.html', {
        'form': form,
        'form_title': 'Create Shop'
    })

@login_required
@user_passes_test(is_admin_or_staff)
def shop_price_create(request):
    if request.method == 'POST':
        form = ShopProductPriceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Shop Price created successfully!')
            return redirect('shop_price_list')
    else:
        form = ShopProductPriceForm()
    return render(request, 'accounts/model_form.html', {
        'form': form,
        'form_title': 'Create Shop Price'
    })